import os
from datetime import datetime

from util.excel_util import init_ws, write
from util.mysql_util import get_columns
import openpyxl.styles as sty


def only_start(only):
    if (only[0:1] == "#"):
        if (only[1:2] != "#"):
            return 1
    return 0


def read_txt(ws, config):
    fs = open("util/analyse.txt", encoding="utf-8")
    lines = fs.readlines()
    unit = []
    row = 2
    init_ws(ws)
    for index, line in enumerate(lines):
        if (index < len(lines) - 1):
            is_start = (only_start(line) & lines[index + 1].startswith("###"))
            if (is_start | line.startswith("###")):
                if (is_start):
                    unit = []
                unit.append(lines[index].replace("\n", ""))
            is_over = (only_start(lines[index + 1]) & lines[index].startswith("###"))
            if (is_over):
                formart_sql(unit, config)
                write_excel(ws, row, unit)
                row += 1
    print("ok,你现在可以到/result/文件夹下找到结果了🌈")

def formart_time(element):
    str_time = element.replace("\n", "")[1:16]
    index = str_time.index(" ")
    d = str_time[index - 2:index]
    m = str_time[index - 4:index - 2]
    y = str_time[0:index - 4]
    t = str_time[index:len(str_time)]
    r = y + "/" + m + "/" + d + "" + t
    return datetime.strptime(r, "%y/%m/%d %H:%M:%S")


def get_type_db_table(element):
    if (element.startswith('### DELETE FROM ')):
        index = element.index(".")
        return "DELETE", element[17:index - 1], element[index + 2:len(element) - 1]
    elif (element.startswith('### INSERT INTO ')):
        index = element.index(".")
        return "INSERT", element[17:index - 1], element[index + 2:len(element) - 1]
    elif (element.startswith('### UPDATE ')):
        index = element.index(".")
        return "UPDATE", element[12:index - 1], element[index + 2:len(element) - 1]


def for_fun(l, sql, index, mid, end, list):
    for i in range(l):
        if (i < l - 1):
            sql += list[index + i] + mid
        else:
            sql += list[index + i] + end
    return sql


def formart_sql(unit, config):
    unit[0] = formart_time(unit[0])
    type, db, table = get_type_db_table(unit[1])
    unit[1] = type
    unit.append(db)
    unit.append(table)
    unit[2] = unit[2].replace("### ", "")
    columns = get_columns(config[0], config[1], config[2], db, table)
    unit.append(len(columns))
    for i, c in enumerate(columns):
        unit[3 + i] = unit[3 + i].replace("###   @" + str(i + 1), c)
    if (type == "UPDATE"):
        for i, c in enumerate(columns):
            unit[len(columns) + 4 + i] = unit[len(columns) + 4 + i].replace("###   @" + str(i + 1), c)
        unit[len(columns) + 3] = unit[len(columns) + 3].replace("### ", "")
    return unit


def write_excel(ws, row, list):
    ll = len(list)
    l = list[ll - 1]
    sql = ''
    back_sql = ''
    fill = sty.PatternFill(fill_type='solid', fgColor="e32636")
    if (list[1] == "DELETE"):
        sql = "DELETE FROM `" + list[ll - 3] + "`.`" + list[ll - 2] + "` WHERE "
        sql = for_fun(l, sql, 3, " AND ", ";", list)
        back_sql = sql.replace("DELETE FROM", "INSERT INTO") \
            .replace("WHERE", "VALUES (") \
            .replace("AND", ",") \
            .replace(";", ");")
    if (list[1] == "INSERT"):
        fill = sty.PatternFill(fill_type='solid', fgColor="4de680")
        sql = "INSERT INTO `" + list[ll - 3] + "`.`" + list[ll - 2] + "` VALUES ("
        sql = for_fun(l, sql, 3, " , ", ");", list)
        back_sql = sql.replace("INSERT INTO", "DELETE FROM") \
            .replace("VALUES (", "WHERE ") \
            .replace(",", "AND") \
            .replace(");", ";")
    if (list[1] == "UPDATE"):
        sql = "UPDATE `" + list[ll - 3] + "`.`" + list[ll - 2] + "` SET "
        sql = for_fun(l, sql, 3, " , ", "", list)
        sql = sql + " WHERE "
        sql = for_fun(l, sql, l + 4, " AND ", ";", list)
        back_sql = "UPDATE `" + list[ll - 3] + "`.`" + list[ll - 2] + "` SET "
        back_sql = for_fun(l, back_sql, l + 4, " , ", "", list)
        back_sql = back_sql + " WHERE "
        back_sql = for_fun(l, back_sql, 3, " AND ", ";", list)

        fill = sty.PatternFill(fill_type='solid', fgColor="f28500")
    ws.cell(row=row, column=1, value='%s' % list[0])
    ws.cell(row=row, column=2, value='%s' % list[1]).fill = fill
    ws.cell(row=row, column=3, value='%s' % list[ll - 3])
    ws.cell(row=row, column=4, value='%s' % list[ll - 2])
    ws.cell(row=row, column=5, value='%s' % back_sql)
    ws.cell(row=row, column=6, value='%s' % sql)


def get_file_name():
    rootdir = 'log0000'
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    dirs = []
    for i in list:
        dirs.append(i)
    return dirs


def get_command(d):
    command = 'mysqlbinlog  --base64-output="decode-rows" -v  log0000/' + d + ' >util/analyse.txt'
    index = d.index(".")
    name = d[index + 1:len(d)]
    return command, name


def startlog(config):
    dir = get_file_name()
    for d in dir:
        command, name = get_command(d)
        os.system(command)
        write(read_txt, name, config)
    print("")
    print("拍拍手上的灰--完事了！🦄")


if __name__ == '__main__':
    write(read_txt, "C:\\Users\\mshu\\Desktop\\log.xlsx")
