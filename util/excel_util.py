import traceback

from openpyxl import Workbook
import openpyxl.styles as sty


def init_ws(ws):
    fill = sty.PatternFill(fill_type='solid', fgColor="20b2aa")
    ws.cell(row=1, column=1, value='时间').fill = fill
    ws.cell(row=1, column=2, value='类型').fill = fill
    ws.cell(row=1, column=3, value='数据库').fill = fill
    ws.cell(row=1, column=4, value='表').fill = fill
    ws.cell(row=1, column=5, value='回滚SQL').fill = fill
    ws.cell(row=1, column=6, value='执行SQL').fill = fill
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 7


def write(fn, path, conf):
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = "mysqlog"
    try:
        fn(ws, conf)
        wb.save(filename="result/" + path + ".xlsx")
        return True
    except Exception as e:
        print(traceback.extract_stack(e))
        print("保存异常 ！！！")
        return False


if __name__ == '__main__':
    print()
